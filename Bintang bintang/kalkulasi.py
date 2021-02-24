{
 "cells": [
  {
   "cell_type": "code",
   "execution_count": 2,
   "metadata": {
    "scrolled": false
   },
   "outputs": [
    {
     "name": "stderr",
     "output_type": "stream",
     "text": [
      "/home/rian-budiansyah/anaconda3/lib/python3.7/site-packages/ipykernel_launcher.py:17: DeprecationWarning: Call to deprecated function get_sheet_by_name (Use wb[sheetname]).\n",
      "/home/rian-budiansyah/anaconda3/lib/python3.7/site-packages/ipykernel_launcher.py:33: DeprecationWarning: Call to deprecated function get_sheet_by_name (Use wb[sheetname]).\n"
     ]
    },
    {
     "name": "stdout",
     "output_type": "stream",
     "text": [
      "alpha ori\n",
      "05h55m10.30536s\n",
      "07h24m25.4304s \n"
     ]
    },
    {
     "ename": "ValueError",
     "evalue": "Latitude angle(s) must be within -90 deg <= angle <= 90 deg, got 111.10595999999998 deg",
     "output_type": "error",
     "traceback": [
      "\u001b[0;31m---------------------------------------------------------------------------\u001b[0m",
      "\u001b[0;31mValueError\u001b[0m                                Traceback (most recent call last)",
      "\u001b[0;32m<ipython-input-2-8a39f3c4de0a>\u001b[0m in \u001b[0;36m<module>\u001b[0;34m\u001b[0m\n\u001b[1;32m     39\u001b[0m \u001b[0mprint\u001b[0m\u001b[0;34m(\u001b[0m\u001b[0mdeclination\u001b[0m\u001b[0;34m)\u001b[0m\u001b[0;34m\u001b[0m\u001b[0;34m\u001b[0m\u001b[0m\n\u001b[1;32m     40\u001b[0m \u001b[0;34m\u001b[0m\u001b[0m\n\u001b[0;32m---> 41\u001b[0;31m \u001b[0mcoordinates\u001b[0m \u001b[0;34m=\u001b[0m \u001b[0mSkyCoord\u001b[0m\u001b[0;34m(\u001b[0m\u001b[0mright_Ascension\u001b[0m\u001b[0;34m,\u001b[0m \u001b[0mdeclination\u001b[0m\u001b[0;34m,\u001b[0m \u001b[0mframe\u001b[0m\u001b[0;34m=\u001b[0m\u001b[0;34m'icrs'\u001b[0m\u001b[0;34m)\u001b[0m\u001b[0;34m\u001b[0m\u001b[0;34m\u001b[0m\u001b[0m\n\u001b[0m\u001b[1;32m     42\u001b[0m \u001b[0mbintang\u001b[0m \u001b[0;34m=\u001b[0m \u001b[0mFixedTarget\u001b[0m\u001b[0;34m(\u001b[0m\u001b[0mname\u001b[0m\u001b[0;34m=\u001b[0m\u001b[0mbintang\u001b[0m\u001b[0;34m,\u001b[0m \u001b[0mcoord\u001b[0m\u001b[0;34m=\u001b[0m\u001b[0mcoordinates\u001b[0m\u001b[0;34m)\u001b[0m\u001b[0;34m\u001b[0m\u001b[0;34m\u001b[0m\u001b[0m\n\u001b[1;32m     43\u001b[0m \u001b[0;34m\u001b[0m\u001b[0m\n",
      "\u001b[0;32m~/anaconda3/lib/python3.7/site-packages/astropy/coordinates/sky_coordinate.py\u001b[0m in \u001b[0;36m__init__\u001b[0;34m(self, copy, *args, **kwargs)\u001b[0m\n\u001b[1;32m    253\u001b[0m             \u001b[0margs\u001b[0m \u001b[0;34m=\u001b[0m \u001b[0mlist\u001b[0m\u001b[0;34m(\u001b[0m\u001b[0margs\u001b[0m\u001b[0;34m)\u001b[0m  \u001b[0;31m# Make it mutable\u001b[0m\u001b[0;34m\u001b[0m\u001b[0;34m\u001b[0m\u001b[0m\n\u001b[1;32m    254\u001b[0m             skycoord_kwargs, components, info = _parse_coordinate_data(\n\u001b[0;32m--> 255\u001b[0;31m                 frame_cls(**frame_kwargs), args, kwargs)\n\u001b[0m\u001b[1;32m    256\u001b[0m \u001b[0;34m\u001b[0m\u001b[0m\n\u001b[1;32m    257\u001b[0m             \u001b[0;31m# In the above two parsing functions, these kwargs were identified\u001b[0m\u001b[0;34m\u001b[0m\u001b[0;34m\u001b[0m\u001b[0;34m\u001b[0m\u001b[0m\n",
      "\u001b[0;32m~/anaconda3/lib/python3.7/site-packages/astropy/coordinates/sky_coordinate_parsers.py\u001b[0m in \u001b[0;36m_parse_coordinate_data\u001b[0;34m(frame, args, kwargs)\u001b[0m\n\u001b[1;32m    299\u001b[0m                                                                   repr_attr_names, units):\n\u001b[1;32m    300\u001b[0m                 \u001b[0mattr_class\u001b[0m \u001b[0;34m=\u001b[0m \u001b[0mframe\u001b[0m\u001b[0;34m.\u001b[0m\u001b[0mrepresentation_type\u001b[0m\u001b[0;34m.\u001b[0m\u001b[0mattr_classes\u001b[0m\u001b[0;34m[\u001b[0m\u001b[0mrepr_attr_name\u001b[0m\u001b[0;34m]\u001b[0m\u001b[0;34m\u001b[0m\u001b[0;34m\u001b[0m\u001b[0m\n\u001b[0;32m--> 301\u001b[0;31m                 \u001b[0m_components\u001b[0m\u001b[0;34m[\u001b[0m\u001b[0mframe_attr_name\u001b[0m\u001b[0;34m]\u001b[0m \u001b[0;34m=\u001b[0m \u001b[0mattr_class\u001b[0m\u001b[0;34m(\u001b[0m\u001b[0marg\u001b[0m\u001b[0;34m,\u001b[0m \u001b[0munit\u001b[0m\u001b[0;34m=\u001b[0m\u001b[0munit\u001b[0m\u001b[0;34m)\u001b[0m\u001b[0;34m\u001b[0m\u001b[0;34m\u001b[0m\u001b[0m\n\u001b[0m\u001b[1;32m    302\u001b[0m \u001b[0;34m\u001b[0m\u001b[0m\n\u001b[1;32m    303\u001b[0m         \u001b[0;32melse\u001b[0m\u001b[0;34m:\u001b[0m\u001b[0;34m\u001b[0m\u001b[0;34m\u001b[0m\u001b[0m\n",
      "\u001b[0;32m~/anaconda3/lib/python3.7/site-packages/astropy/coordinates/angles.py\u001b[0m in \u001b[0;36m__new__\u001b[0;34m(cls, angle, unit, **kwargs)\u001b[0m\n\u001b[1;32m    505\u001b[0m             \u001b[0;32mraise\u001b[0m \u001b[0mTypeError\u001b[0m\u001b[0;34m(\u001b[0m\u001b[0;34m\"A Latitude angle cannot be created from a Longitude angle\"\u001b[0m\u001b[0;34m)\u001b[0m\u001b[0;34m\u001b[0m\u001b[0;34m\u001b[0m\u001b[0m\n\u001b[1;32m    506\u001b[0m         \u001b[0mself\u001b[0m \u001b[0;34m=\u001b[0m \u001b[0msuper\u001b[0m\u001b[0;34m(\u001b[0m\u001b[0;34m)\u001b[0m\u001b[0;34m.\u001b[0m\u001b[0m__new__\u001b[0m\u001b[0;34m(\u001b[0m\u001b[0mcls\u001b[0m\u001b[0;34m,\u001b[0m \u001b[0mangle\u001b[0m\u001b[0;34m,\u001b[0m \u001b[0munit\u001b[0m\u001b[0;34m=\u001b[0m\u001b[0munit\u001b[0m\u001b[0;34m,\u001b[0m \u001b[0;34m**\u001b[0m\u001b[0mkwargs\u001b[0m\u001b[0;34m)\u001b[0m\u001b[0;34m\u001b[0m\u001b[0;34m\u001b[0m\u001b[0m\n\u001b[0;32m--> 507\u001b[0;31m         \u001b[0mself\u001b[0m\u001b[0;34m.\u001b[0m\u001b[0m_validate_angles\u001b[0m\u001b[0;34m(\u001b[0m\u001b[0;34m)\u001b[0m\u001b[0;34m\u001b[0m\u001b[0;34m\u001b[0m\u001b[0m\n\u001b[0m\u001b[1;32m    508\u001b[0m         \u001b[0;32mreturn\u001b[0m \u001b[0mself\u001b[0m\u001b[0;34m\u001b[0m\u001b[0;34m\u001b[0m\u001b[0m\n\u001b[1;32m    509\u001b[0m \u001b[0;34m\u001b[0m\u001b[0m\n",
      "\u001b[0;32m~/anaconda3/lib/python3.7/site-packages/astropy/coordinates/angles.py\u001b[0m in \u001b[0;36m_validate_angles\u001b[0;34m(self, angles)\u001b[0m\n\u001b[1;32m    522\u001b[0m         \u001b[0;32mif\u001b[0m \u001b[0mnp\u001b[0m\u001b[0;34m.\u001b[0m\u001b[0many\u001b[0m\u001b[0;34m(\u001b[0m\u001b[0mangles\u001b[0m\u001b[0;34m.\u001b[0m\u001b[0mvalue\u001b[0m \u001b[0;34m<\u001b[0m \u001b[0mlower\u001b[0m\u001b[0;34m)\u001b[0m \u001b[0;32mor\u001b[0m \u001b[0mnp\u001b[0m\u001b[0;34m.\u001b[0m\u001b[0many\u001b[0m\u001b[0;34m(\u001b[0m\u001b[0mangles\u001b[0m\u001b[0;34m.\u001b[0m\u001b[0mvalue\u001b[0m \u001b[0;34m>\u001b[0m \u001b[0mupper\u001b[0m\u001b[0;34m)\u001b[0m\u001b[0;34m:\u001b[0m\u001b[0;34m\u001b[0m\u001b[0;34m\u001b[0m\u001b[0m\n\u001b[1;32m    523\u001b[0m             raise ValueError('Latitude angle(s) must be within -90 deg <= angle <= 90 deg, '\n\u001b[0;32m--> 524\u001b[0;31m                              'got {}'.format(angles.to(u.degree)))\n\u001b[0m\u001b[1;32m    525\u001b[0m \u001b[0;34m\u001b[0m\u001b[0m\n\u001b[1;32m    526\u001b[0m     \u001b[0;32mdef\u001b[0m \u001b[0m__setitem__\u001b[0m\u001b[0;34m(\u001b[0m\u001b[0mself\u001b[0m\u001b[0;34m,\u001b[0m \u001b[0mitem\u001b[0m\u001b[0;34m,\u001b[0m \u001b[0mvalue\u001b[0m\u001b[0;34m)\u001b[0m\u001b[0;34m:\u001b[0m\u001b[0;34m\u001b[0m\u001b[0;34m\u001b[0m\u001b[0m\n",
      "\u001b[0;31mValueError\u001b[0m: Latitude angle(s) must be within -90 deg <= angle <= 90 deg, got 111.10595999999998 deg"
     ]
    }
   ],
   "source": [
    "\"\"\"\n",
    "Created on januari\n",
    "\n",
    "@author: Rian Budiansyah\n",
    "\"\"\"\n",
    "\n",
    "import astropy.units as u\n",
    "from astropy.coordinates import EarthLocation, SkyCoord\n",
    "import matplotlib.pyplot as plt\n",
    "from pytz import timezone\n",
    "\n",
    "from astroplan import Observer, FixedTarget\n",
    "from astroplan.plots import plot_airmass\n",
    "import openpyxl\n",
    "\n",
    "excel_document = openpyxl.load_workbook('databasetelescope.xlsx')\n",
    "sheet = excel_document.get_sheet_by_name('Sheet1')\n",
    "#excel_document.sheetnames('Sheet1')\n",
    "latitude = sheet['D68'].value\n",
    "longitude = sheet['E68'].value\n",
    "elevation = sheet['F68'].value * u.m\n",
    "\n",
    "location = EarthLocation.from_geodetic(longitude, latitude, elevation)\n",
    "\n",
    "################# perlu penyesuaian di penempatan nama, timezone dan descripsi pada daerah ini #########\n",
    "bosscha = Observer(name= 'Zeiss Telescope',\n",
    "                  location=location,\n",
    "                  timezone=timezone('Etc/GMT+7'),\n",
    "                  description=\"Bosscha Observatory\")\n",
    "###########################################################################################################\n",
    "\n",
    "excel_document_two = openpyxl.load_workbook('koordinat.xlsx')\n",
    "sheet2 = excel_document_two.get_sheet_by_name('Sheet1')\n",
    "bintang = sheet2['B94'].value\n",
    "right_Ascension = sheet2['C94'].value\n",
    "declination = sheet2['D94'].value\n",
    "print(bintang)\n",
    "print(right_Ascension)\n",
    "print(declination)\n",
    "\n",
    "coordinates = SkyCoord(right_Ascension, declination, frame='icrs')\n",
    "bintang = FixedTarget(name=bintang, coord=coordinates)\n",
    "\n",
    "from astropy.time import Time\n",
    "\n",
    "time = Time('2019-12-29 12:00:00') #WAKTU LAPTOP SEHARUSNYA\n",
    "\n",
    "plot_airmass(bintang, bosscha, time)\n",
    "\n",
    "ax = plt.gca()\n",
    "box = ax.get_position()\n",
    "ax.set_position([box.x0, box.y0, box.width * 0.8, box.height * 0.8])\n",
    "\n",
    "plt.legend(loc=1, bbox_to_anchor=(1.35, 1))\n",
    "plt.tight_layout()\n",
    "plt.show()\n"
   ]
  },
  {
   "cell_type": "code",
   "execution_count": null,
   "metadata": {},
   "outputs": [],
   "source": []
  },
  {
   "cell_type": "code",
   "execution_count": null,
   "metadata": {},
   "outputs": [],
   "source": []
  },
  {
   "cell_type": "code",
   "execution_count": null,
   "metadata": {},
   "outputs": [],
   "source": []
  }
 ],
 "metadata": {
  "kernelspec": {
   "display_name": "Python 3",
   "language": "python",
   "name": "python3"
  },
  "language_info": {
   "codemirror_mode": {
    "name": "ipython",
    "version": 3
   },
   "file_extension": ".py",
   "mimetype": "text/x-python",
   "name": "python",
   "nbconvert_exporter": "python",
   "pygments_lexer": "ipython3",
   "version": "3.7.6"
  }
 },
 "nbformat": 4,
 "nbformat_minor": 2
}
